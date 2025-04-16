import os
import pandas as pd
from io import StringIO
import openpyxl

def process_fluid_data():
    folder_path = 'E:/RIGHT MODEL/Injection_6000_v35'
    output_excel = os.path.join(folder_path, 'output_data.xlsx')

    # 第一阶段：文件处理（保持不变）
    # ... [原有文件处理逻辑] ...

    # 第二阶段：平均值计算（修改部分）
    try:
        wb = openpyxl.load_workbook(output_excel)
        ws = wb.active

        # 寻找起始行（时间 > 10）
        start_row = None
        for row in range(1, ws.max_row + 1):
            time_cell = ws.cell(row=row, column=1)
            if time_cell.value and float(time_cell.value) > 10:
                start_row = row
                break

        if not start_row:
            raise ValueError("未找到时间>10的有效起始行")

        # 计算所有有效行的D值
        total = 0.0
        valid_rows = 0
        valid_data = []

        # 修改点1：移除行数限制
        for row in range(start_row, ws.max_row + 1):
            # 获取B、C列值
            try:
                b = abs(float(ws.cell(row=row, column=2).value))
                c = abs(float(ws.cell(row=row, column=3).value))
            except (TypeError, ValueError, AttributeError):
                continue  # 跳过无效数据

            d = round(b * c, 7)
            total += d
            valid_rows += 1
            valid_data.append( (row, d) )  # 记录所有有效数据

        if valid_rows == 0:
            raise ValueError("没有找到有效数据")

        # 生成最终报告
        print(f"\n分析报告 ({output_excel})")
        print("=" * 60)
        print(f"稳态起始行: {start_row}")
        print(f"总有效行数: {valid_rows} (从{start_row}行到{ws.max_row}行)")
        
        # 显示前20行示例（如果存在）
        if valid_data:
            print("\n前20行D值示例:")
            for idx, (row_num, d) in enumerate(valid_data[:20], 1):
                print(f"行{row_num}: {d:.7f}")

        print("\n最终统计:")
        print(f"D列平均值（共{valid_rows}行）: {total/valid_rows:.7f}")
        print("=" * 60)

    except Exception as e:
        print(f"处理失败: {str(e)}")
        exit(1)

if __name__ == "__main__":
    process_fluid_data()