import openpyxl
import sys

def process_excel_data(input_file):
    """处理Excel文件并返回格式化结果"""
    try:
        wb = openpyxl.load_workbook(input_file)
    except FileNotFoundError:
        raise ValueError(f"文件不存在: {input_file}")
    except Exception as e:
        raise ValueError(f"文件读取失败: {str(e)}")
    
    sheet = wb.active

    # 寻找第一个A列有效行
    start_row = None
    for row in range(1, sheet.max_row + 1):
        a_cell = sheet.cell(row=row, column=1)
        if a_cell.value is None:
            continue
            
        try:
            a_val = abs(float(a_cell.value))
        except (TypeError, ValueError):
            continue
            
        if a_val > 10:
            start_row = row
            break

    if start_row is None:
        raise ValueError("未找到A列值大于10的行")

    # 收集数据和计算
    data_samples = []
    current_row = start_row
    total = 0.0
    count = 0

    while current_row <= sheet.max_row and len(data_samples) < 20:
        # 获取三列数据
        a_cell = sheet.cell(row=current_row, column=1)
        b_cell = sheet.cell(row=current_row, column=2)
        c_cell = sheet.cell(row=current_row, column=3)

        # 验证A列
        try:
            a_val = abs(float(a_cell.value)) if a_cell.value else None
        except:
            a_val = None
            
        if a_val is None or a_val <= 10:
            current_row += 1
            continue

        # 处理B、C列
        try:
            b = abs(float(b_cell.value)) if b_cell.value else None
            c = abs(float(c_cell.value)) if c_cell.value else None
        except:
            current_row += 1
            continue

        if None in (b, c):
            current_row += 1
            continue

        # 计算D值
        d = abs(b * c)  # 确保绝对値
        total += d
        data_samples.append( (current_row, b, c, d) )
        count += 1
        current_row += 1

    if count == 0:
        raise ValueError("没有有效数据可供计算")

    return total/count, start_row, data_samples

if __name__ == "__main__":
    DEFAULT_PATH = r"E:/NEW_Grid/Injection_10000_V15/output_data.xlsx"
    file_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH

    try:
        average, start_row, samples = process_excel_data(file_path)
        
        # 打印报告
        print("="*60)
        print(f"文件路径: {file_path}")
        print(f"首个有效行号: {start_row}")
        print("\n前10行数据示例 (B/C/D列):")
        
        # 输出前10行完整数据
        for idx, (row, b, c, d) in enumerate(samples[:10], 1):
            print(f"行{row}: B={b:.7f}  C={c:.7f}  =>  D={d:.7f}")
        
        # 输出剩余行简要信息
        print("\n后续有效行数据:")
        for row, _, _, d in samples[10:]:
            print(f"行{row}: D={d:.7f}")
        
        print("\n统计结果:")
        print(f"平均值 (D列，7位小数): {average:.7f}")
        print("="*60)
        
    except Exception as e:
        print(f"处理错误: {str(e)}")
        sys.exit(1)