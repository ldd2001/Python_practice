import os
import pandas as pd
from io import StringIO
import openpyxl

folder_path = 'E:/NEW CFB MODEL/Injection_4000_v25'  # 文件夹路径
output_excel_path = folder_path + '/output_data.xlsx'  # 输出 Excel 文件路径

# 定义函数删除文件的前n行
def modify_file(file_path, lines_to_delete):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()

        # 删除前n行
        modified_lines = lines[lines_to_delete:]

        # 返回修改后的内容
        return modified_lines

    except Exception as e:
        print(f"修改文件时出错: {e} 文件路径: {file_path}")       
        exit()

# 定义函数从文件中提取数据并输出到 Excel
def extract_and_export_to_excel(folder_path, output_excel_path):
    try:
        # 在指定文件夹内精确匹配文件名
        file_1, file_2, file_3 = None, None, None
        for file_name in os.listdir(folder_path):
            if file_name == 'FLUX_OUTLET':
                file_1 = os.path.join(folder_path, file_name)
            elif file_name == 'FLUX_OUTLET_fluidSpecies':
                file_2 = os.path.join(folder_path, file_name)
            elif file_name == 'FLUX_INLET':
                file_3 = os.path.join(folder_path, file_name)

        # 确保找到了所有必要的文件
        if not file_1:
            raise FileNotFoundError(f"在文件夹 {folder_path} 中没有找到 FLUX_OUTLET 文件。")
        if not file_2:
            raise FileNotFoundError(f"在文件夹 {folder_path} 中没有找到 FLUX_OUTLET_fluidSpecies 文件。")
        if not file_3:
            raise FileNotFoundError(f"在文件夹 {folder_path} 中没有找到 FLUX_INLET 文件。")

        print(f"找到 FLUX_OUTLET 文件: {file_1}")
        print(f"找到 FLUX_OUTLET_fluidSpecies 文件: {file_2}")
        print(f"找到 FLUX_INLET 文件: {file_3}\n")

        # 删除 FLUX_OUTLET 文件的前23行
        modified_file_1 = modify_file(file_1, 23)
        # 删除 FLUX_OUTLET_fluidSpecies 文件的前11行
        modified_file_2 = modify_file(file_2, 11)
        # 删除 FLUX_INLET 文件的前23行
        modified_file_3 = modify_file(file_3, 23)

        # 将文件内容转为 DataFrame，处理错误的行
        try:
            df_1 = pd.read_csv(StringIO(''.join(modified_file_1)), header=None, sep=r'\s+', on_bad_lines='skip')
            print(f"成功读取 {file_1}")
        except Exception as e:
            print(f"读取 {file_1} 时出错: {e}")
            raise

        try:
            df_2 = pd.read_csv(StringIO(''.join(modified_file_2)), header=None, sep=r'\s+', on_bad_lines='skip')
            print(f"成功读取 {file_2}")
        except Exception as e:
            print(f"读取 {file_2} 时出错: {e}")
            raise

        try:
            df_3 = pd.read_csv(StringIO(''.join(modified_file_3)), header=None, sep=r'\s+', on_bad_lines='skip')
            print(f"成功读取 {file_3}\n")
        except Exception as e:
            print(f"读取 {file_3} 时出错: {e}")
            raise

        # 检查文件是否有足够的行和列
        if df_1.shape[1] < 2:
            raise ValueError(f"文件 {file_1} 没有足够的列")
        if df_2.shape[1] < 5:
            raise ValueError(f"文件 {file_2} 没有足够的列")
        if df_3.shape[1] < 2:
            raise ValueError(f"文件 {file_3} 没有足够的列")

        # 提取 FLUX_OUTLET 的前两列：时间与出口质量流量
        data_1 = df_1.iloc[:, :2]

        # 提取 FLUX_OUTLET_fluidSpecies 的第五列：出口SO2质量分数
        data_2 = df_2.iloc[:, 4]

        # 提取 FLUX_INLET 的第二列（第24行数据）：进口质量流量
        inlet_value = df_3.iloc[0, 1]  # 由于前23行已删除，所以第24行对应的索引是0

        # 将两个数据合并并转换为浮动数据类型
        final_data = pd.concat([data_1, data_2], axis=1)

        # 将 final_data 中的所有列转换为 float 类型
        final_data = final_data.apply(pd.to_numeric, errors='coerce')

        # 将结果保存为 Excel 文件
        final_data.to_excel(output_excel_path, index=False)
        print(f"抽取数据并导出到 Excel 文件: {output_excel_path}")

        # 打开并修改 Excel 文件
        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb.active
        ws.delete_rows(1)  # 删除第一行

        # 设置前三列的单元格格式为数字，小数点后七位
        for col in range(1, 4):  # 针对前三列
            for row in range(1, ws.max_row + 1):  # 针对每一行
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:  # 确保单元格有值
                    cell.number_format = '0.0000000'  # 设置为数字格式，保留7位小数

        # 自动调整列宽，以确保所有数字能够正确显示
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = openpyxl.utils.get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell_value = str(ws[f"{column}{row}"].value)
                max_length = max(max_length, len(cell_value))
            adjusted_width = (max_length + 2)  # 增加一些空间
            ws.column_dimensions[column].width = adjusted_width

        # 保存修改后的 Excel 文件
        wb.save(output_excel_path)
        print(f"Excel 文件格式已更新并保存: {output_excel_path}\n")


        # 处理检索功能：检索第一列，当第一列的数值大于等于10时，计算第二列和第三列的平均值
        #print("正在检索并计算10s后第二列、第三列的平均值:\n")
        df = pd.read_excel(output_excel_path)  # 读取修改后的 Excel 文件

        # 输出 FLUX_INLET 文件第24行的第二列数据
        print(f"FLUX_INLET 文件第24行的第二列数据（烟气进口质量流量）: {float(inlet_value):.7f} kg/s\n")

        found = False

        for index, row in df.iterrows():
            if row[0] >= 10:  # 如果第一列的值大于等于 10
                # 获取该行及以下所有行的第二列和第三列的平均值
                avg_col2 = df.iloc[index:, 1].mean()  # 第二列的平均值
                avg_col3 = df.iloc[index:, 2].mean()  # 第三列的平均值
                print(f"稳态时间设置为10s")
                print(f"在新的表格中，从第 {index + 2} 行开始，时间大于等于10s")
                print("开始计算10s后第二列、第三列的平均值:\n")
                print(f"第二列的平均值（平均烟气出口质量流量）: {avg_col2:.7f} kg/s")
                print(f"第三列的平均值(SO2平均质量分数): {avg_col3:.7f}\n")
                found = True
                break  # 找到第一个满足条件的行后就退出循环
        if not found:
            max_value = df.iloc[:, 0].max()
            print(f"\nError:读取到的文件未记录10s后的数据，已记录的最大时间为为：{max_value:.7f}s\n")



    except Exception as e:
        print(f"导出到 Excel 时出错: {e}")
        exit()

# 执行提取和导出操作
extract_and_export_to_excel(folder_path, output_excel_path)
